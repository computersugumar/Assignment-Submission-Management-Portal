import os
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime

# Exports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.units import mm

# ---------------------------
# CONFIG: changeable passwords & logo paths
# ---------------------------
LOGIN_PASSWORD = "Admin123"       # change login password here
EXPORT_PASSWORD = "Admin123"     # change export password here

# Local logo paths (change to your real local paths if needed)
COLLEGE_LOGO_PATH = r"E:\IGNOU SMART ASSIGNMENT PORTAL\BHC LOGO.png"
IGNOU_LOGO_PATH = r"E:\IGNOU SMART ASSIGNMENT PORTAL\IGNOU logo.png"

DB_PATH = "assignments.db"

# ---------------------------
# DB initialization
# ---------------------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS assignments (
            RecordNo INTEGER PRIMARY KEY AUTOINCREMENT,
            EnrollmentNo TEXT NOT NULL,
            StudentName TEXT NOT NULL,
            ProgramCode TEXT NOT NULL,
            CourseCode TEXT NOT NULL,
            SubmissionDate TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# ---------------------------
# Helper utilities
# ---------------------------
def center_window(win, w, h):
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = (sw - w) // 2
    y = (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")


# ---------------------------
# Main App Class
# ---------------------------
class IGNOUApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IGNOU Assignment Submission")
        center_window(self, 980, 640)
        self.resizable(True,True)

        # Shared header/footer frames
        self.header_frame = ttk.Frame(self)
        self.header_frame.pack(fill=tk.X)
        self.body_frame = ttk.Frame(self)
        self.body_frame.pack(fill=tk.BOTH, expand=True)
        self.footer_frame = ttk.Frame(self)
        self.footer_frame.pack(fill=tk.X)

        self.create_header()
        self.create_footer()

        # Container for pages
        self.pages = {}
        for P in (LoginPage, WelcomePage, MainMenuPage, EnterDetailsPage, ViewDetailsPage, ExportPage, DeleteDetailPage):
            page = P(parent=self.body_frame, controller=self)
            self.pages[P.__name__] = page
            page.grid(row=0, column=0, sticky='nsew')

        self.show_page('LoginPage')

    def create_header(self):
        # Header: left logo, center title, right logo
        for w in self.header_frame.winfo_children():
            w.destroy()

        left = ttk.Label(self.header_frame)
        left.pack(side=tk.LEFT, padx=10, pady=6)
        center = ttk.Label(self.header_frame, text="IGNOU Assignment Submission", font=("Times New Roman", 22, 'bold',))
        center.pack(side=tk.LEFT, expand=True)
        right = ttk.Label(self.header_frame)
        right.pack(side=tk.RIGHT, padx=10, pady=6)

        # Try loading logos if available
        try:
            if os.path.exists(COLLEGE_LOGO_PATH):
                from PIL import Image, ImageTk
                img = Image.open(COLLEGE_LOGO_PATH)
                img = img.resize((70, 70), Image.LANCZOS)
                left.img = ImageTk.PhotoImage(img)
                left.config(image=left.img)
        except Exception:
            pass
        try:
            if os.path.exists(IGNOU_LOGO_PATH):
                from PIL import Image, ImageTk
                img2 = Image.open(IGNOU_LOGO_PATH)
                img2 = img2.resize((70, 70), Image.LANCZOS)
                right.img = ImageTk.PhotoImage(img2)
                right.config(image=right.img)
        except Exception:
            pass

    def create_footer(self):
        for w in self.footer_frame.winfo_children():
            w.destroy()
        footer_label = ttk.Label(self.footer_frame, text="Developed by Computer Sugumar", font=("Constantia", 16,"bold"),foreground="red")
        footer_label_1 = ttk.Label(self.footer_frame, text="Contact: +91 9344 7344 23", font=("Calibri", 14,"bold"),foreground="blue")
        footer_label_2 = ttk.Label(self.footer_frame, text="E-mail: computersugumarpro@gmail.com", font=("Constantia", 16,"bold"),foreground="green")

        footer_label.pack(pady=0)
        footer_label_1.pack(pady=0)
        footer_label_2.pack(pady=0)
    def show_page(self, page_name):
        page = self.pages[page_name]
        page.tkraise()


# ---------------------------
# Hover effect helper
# ---------------------------
def add_hover(widget, hover_bg='#d9d9d9'):
    # For ttk.Button we simulate via bindings and state changes
    def on_enter(e):
        widget.config(cursor='hand2')
        try:
            widget.state(['active'])
        except Exception:
            pass
    def on_leave(e):
        widget.config(cursor='')
        try:
            widget.state(['!active'])
        except Exception:
            pass
    widget.bind('<Enter>', on_enter)
    widget.bind('<Leave>', on_leave)


# ---------------------------
# Individual Pages
# ---------------------------
class LoginPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        lbl = ttk.Label(self, text="Login", font=("Times New Roman", 24))
        lbl.pack(pady=40)

        frm = ttk.Frame(self)
        frm.pack()
        ttk.Label(frm, text="Password:", font=("Times New Roman", 18)).grid(row=0, column=0, padx=6, pady=6)
        self.pwd_var = tk.StringVar()
        ent = ttk.Entry(frm, textvariable=self.pwd_var, show='*', font=("Times New Roman", 16))
        ent.grid(row=0, column=1, padx=6, pady=6)
        ent.bind('<Return>', lambda e: self.check_login())

        btn = ttk.Button(self, text="Go", command=self.check_login)
        btn.pack(pady=18)
        add_hover(btn)

    def check_login(self):
        if self.pwd_var.get() == LOGIN_PASSWORD:
            self.controller.show_page('WelcomePage')
        else:
            messagebox.showerror("Login Failed", "Incorrect password")


class WelcomePage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        lbl = ttk.Label(self, text="Welcome to the IGNOU Assignment Submission Portal", font=("Times New Roman", 18))
        lbl.pack(pady=40)
        btn = ttk.Button(self, text="Go to Main Menu", command=lambda: controller.show_page('MainMenuPage'))
        btn.pack()
        add_hover(btn)


class MainMenuPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        ttk.Label(self, text="Main Menu", font=("Times New Roman", 20)).pack(pady=20)
        btns = [
            ("Enter Details", 'EnterDetailsPage'),
            ("Show Details", 'ViewDetailsPage'),
            ("Export Details", 'ExportPage'),
            ("Delete Detail", 'DeleteDetailPage')
        ]
        for txt, page in btns:
            b = ttk.Button(self, text=txt, command=lambda p=page: controller.show_page(p))
            b.pack(padx=8, pady=8, ipadx=10, ipady=6)
            add_hover(b)


class EnterDetailsPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        ttk.Label(self, text="Enter Details", font=("Times New Roman", 20)).pack(pady=10)
        form = ttk.Frame(self)
        form.pack(padx=20, pady=6)

        labels = ["Enrollment No", "Name", "Program Code", "Course Code", "Submission Date"]
        self.vars = {}
        for i, lbl in enumerate(labels):
            ttk.Label(form, text=lbl+":", font=("Times New Roman", 14)).grid(row=i, column=0, sticky='e', padx=6, pady=6)
            if lbl == "Submission Date":
                v = tk.StringVar()
                de = DateEntry(form, textvariable=v, date_pattern='dd-MM-yyyy', font=("Times New Roman", 12))
                de.grid(row=i, column=1, padx=6, pady=6)
                self.vars['SubmissionDate'] = v
            else:
                v = tk.StringVar()
                ent = ttk.Entry(form, textvariable=v, font=("Times New Roman", 12), width=30)
                ent.grid(row=i, column=1, padx=6, pady=6)
                key = lbl.replace(' ', '')
                self.vars[key] = v

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=12)
        submit_btn = ttk.Button(btn_frame, text="Submit", command=self.submit)
        submit_btn.grid(row=0, column=0, padx=6)
        add_hover(submit_btn)

        back_btn = ttk.Button(btn_frame, text="Back to Main Menu", command=lambda: controller.show_page('MainMenuPage'))
        back_btn.grid(row=0, column=1, padx=6)
        add_hover(back_btn)

    def submit(self):
        # Validate
        vals = {k: v.get().strip() for k, v in self.vars.items()}
        required_keys = ['EnrollmentNo', 'Name', 'ProgramCode', 'CourseCode', 'SubmissionDate']
        for k in required_keys:
            if not vals.get(k):
                messagebox.showwarning("Validation Error", f"{k} is required")
                return
        # Insert
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('''INSERT INTO assignments (EnrollmentNo, StudentName, ProgramCode, CourseCode, SubmissionDate)
                       VALUES (?, ?, ?, ?, ?)''', (vals['EnrollmentNo'], vals['Name'], vals['ProgramCode'], vals['CourseCode'], vals['SubmissionDate']))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", "Record inserted successfully")
        # Clear
        for v in self.vars.values():
            v.set('')


# ---------------------------
# View Details Page (Updated Grid View)
# ---------------------------
from tkinter import ttk

class ViewDetailsPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        ttk.Label(self, text="View Records", font=("Times New Roman", 20, "bold")).pack(pady=10)

        # Search Frame
        search_frame = ttk.Frame(self)
        search_frame.pack(pady=6, padx=10, fill="x")

        ttk.Label(search_frame, text="Enrollment No:", font=("Times New Roman", 12)).grid(row=0, column=0, padx=6, pady=4)
        self.en_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.en_var, width=18).grid(row=0, column=1, padx=6, pady=4)

        ttk.Label(search_frame, text="Program Code:", font=("Times New Roman", 12)).grid(row=0, column=2, padx=6, pady=4)
        self.pg_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.pg_var, width=12).grid(row=0, column=3, padx=6, pady=4)

        ttk.Label(search_frame, text="Course Code:", font=("Times New Roman", 12)).grid(row=0, column=4, padx=6, pady=4)
        self.cc_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.cc_var, width=12).grid(row=0, column=5, padx=6, pady=4)

        # Button Frame
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=6)
        ttk.Button(btn_frame, text="Show All", command=self.show_all).grid(row=0, column=0, padx=6)
        ttk.Button(btn_frame, text="Search", command=self.search).grid(row=0, column=1, padx=6)
        ttk.Button(btn_frame, text="Reset", command=self.reset).grid(row=0, column=2, padx=6)
        back_btn = ttk.Button(btn_frame, text="Back to Main Menu", command=lambda: controller.show_page('MainMenuPage'))
        back_btn.grid(row=0, column=3, padx=6)
        for w in btn_frame.winfo_children():
            add_hover(w)

        # Treeview Frame (Grid View)
        tree_frame = tk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Style for grid lines
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview",
                        background="#ffffff",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="#ffffff",
                        font=("Times New Roman", 12))
        style.map('Treeview', background=[('selected', '#ececec')])
        style.configure("Treeview.Heading", font=("Times New Roman", 12, "bold"))
        style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # Columns
        cols = ("SNo", "RecordNo", "EnrollmentNo", "StudentName", "ProgramCode", "CourseCode", "SubmissionDate")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings')

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        # Heading & Column Widths
        for c in cols:
            self.tree.heading(c, text=c)
        self.tree.column("SNo", width=60, anchor="center")
        self.tree.column("RecordNo", width=100, anchor="center")
        self.tree.column("EnrollmentNo", width=120, anchor="center")
        self.tree.column("StudentName", width=180, anchor="w")
        self.tree.column("ProgramCode", width=120, anchor="center")
        self.tree.column("CourseCode", width=120, anchor="center")
        self.tree.column("SubmissionDate", width=140, anchor="center")

    def _populate_tree(self, rows):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for idx, row in enumerate(rows, start=1):
            self.tree.insert('', 'end', values=(idx, row[0], row[1], row[2], row[3], row[4], row[5]))

    def show_all(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('SELECT RecordNo, EnrollmentNo, StudentName, ProgramCode, CourseCode, SubmissionDate FROM assignments ORDER BY RecordNo')
        rows = cur.fetchall()
        conn.close()
        self._populate_tree(rows)

    def search(self):
        en = self.en_var.get().strip()
        pg = self.pg_var.get().strip()
        cc = self.cc_var.get().strip()
        query = 'SELECT RecordNo, EnrollmentNo, StudentName, ProgramCode, CourseCode, SubmissionDate FROM assignments'
        clauses = []
        params = []
        if en:
            clauses.append('EnrollmentNo = ?')
            params.append(en)
        if pg:
            clauses.append('ProgramCode = ?')
            params.append(pg)
        if cc:
            clauses.append('CourseCode = ?')
            params.append(cc)
        if clauses:
            query += ' WHERE ' + ' AND '.join(clauses)
        query += ' ORDER BY RecordNo'
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        conn.close()
        self._populate_tree(rows)

    def reset(self):
        self.en_var.set('')
        self.pg_var.set('')
        self.cc_var.set('')
        for i in self.tree.get_children():
            self.tree.delete(i)
class ExportPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        ttk.Label(self, text="Export Details ", font=("Times New Roman", 18)).pack(pady=10)
        frm = ttk.Frame(self)
        frm.pack(pady=6)
        ttk.Label(frm, text="Password:", font=("Times New Roman", 12)).grid(row=0, column=0, padx=6)
        self.pvar = tk.StringVar()
        ent = ttk.Entry(frm, textvariable=self.pvar, show='*')
        ent.grid(row=0, column=1, padx=6)
        ent.bind('<Return>', lambda e: self.validate())
        btn = ttk.Button(frm, text="Validate", command=self.validate)
        btn.grid(row=0, column=2, padx=6)
        add_hover(btn)

        self.export_frame = ttk.Frame(self)
        self.export_frame.pack(pady=10)

        back_btn = ttk.Button(self, text="Back to Main Menu", command=lambda: controller.show_page('MainMenuPage'))
        back_btn.pack(pady=4)
        add_hover(back_btn)

    def validate(self):
        if self.pvar.get() == EXPORT_PASSWORD:
            for w in self.export_frame.winfo_children():
                w.destroy()
            b1 = ttk.Button(self.export_frame, text="Export as Excel", command=self.export_excel)
            b1.pack(side=tk.LEFT, padx=8)
            add_hover(b1)
            """b2 = ttk.Button(self.export_frame, text="Export as PDF", command=self.export_pdf)
            b2.pack(side=tk.LEFT, padx=8)
            add_hover(b2)"""
        else:
            messagebox.showerror("Error", "Incorrect export password")

    def _fetch_all(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('SELECT RecordNo, EnrollmentNo, StudentName, ProgramCode, CourseCode, SubmissionDate FROM assignments ORDER BY RecordNo')
        rows = cur.fetchall()
        conn.close()
        return rows

    def export_excel(self):
        rows = self._fetch_all()
        if not rows:
            messagebox.showinfo("No Data", "No records to export")
            return
        path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel Files', '*.xlsx')])
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        # Header row (custom header: college + title + date)
        ws.merge_cells('A1:F1')
        header_text = 'IGNOU Assignment Submission - Exported on ' + datetime.now().strftime('%d-%m-%Y %H:%M:%S')
        ws['A1'] = header_text
        ws['A1'].font = Font(name='Times New Roman', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        # Column headers
        cols = ['S.No', 'RecordNo', 'EnrollmentNo', 'StudentName', 'ProgramCode', 'CourseCode', 'SubmissionDate']
        for cidx, c in enumerate(cols, start=1):
            ws.cell(row=2, column=cidx, value=c)
            ws.cell(row=2, column=cidx).font = Font(name='Times New Roman', size=14, bold=True)
        # Data
        for idx, row in enumerate(rows, start=1):
            rec = [idx, row[0], row[1], row[2], row[3], row[4], row[5]]
            for cidx, val in enumerate(rec, start=1):
                ws.cell(row=2+idx, column=cidx, value=val)
                ws.cell(row=2+idx, column=cidx).font = Font(name='Times New Roman', size=14)
        # Footer
        footer_row = 5+ len(rows)
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=7)
        ws.cell(row=footer_row, column=1, value='Developed by Computer Sugumar').font = Font(name='Times New Roman', size=12)
        try:
            wb.save(path)
            messagebox.showinfo('Exported', f'Excel saved to {path}')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to save Excel: {e}')

class DeleteDetailPage(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        ttk.Label(self, text="Delete Detail", font=("Times New Roman", 20)).pack(pady=10)
        frm = ttk.Frame(self)
        frm.pack(pady=8)
        ttk.Label(frm, text="Enter Record Number to Delete:", font=("Times New Roman", 12)).grid(row=0, column=0, padx=6)
        self.rec_var = tk.StringVar()
        ent = ttk.Entry(frm, textvariable=self.rec_var)
        ent.grid(row=0, column=1, padx=6)
        del_btn = ttk.Button(frm, text="Delete", command=self.delete_record)
        del_btn.grid(row=0, column=2, padx=6)
        add_hover(del_btn)
        back_btn = ttk.Button(frm, text="Back to Main Menu", command=lambda: controller.show_page('MainMenuPage'))
        back_btn.grid(row=0, column=3, padx=6)
        add_hover(back_btn)

        # Also show a small preview tree
        cols = ("SNo", "RecordNo", "EnrollmentNo", "StudentName", "ProgramCode", "CourseCode", "SubmissionDate")
        self.tree = ttk.Treeview(self, columns=cols, show='headings', height=8)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, anchor='center')
        self.tree.pack(padx=10, pady=8, fill=tk.X)
        ttk.Button(self, text="Refresh Preview", command=self.refresh_preview).pack()

    def refresh_preview(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('SELECT RecordNo, EnrollmentNo, StudentName, ProgramCode, CourseCode, SubmissionDate FROM assignments ORDER BY RecordNo')
        rows = cur.fetchall()
        conn.close()
        for i in self.tree.get_children():
            self.tree.delete(i)
        for idx, row in enumerate(rows, start=1):
            self.tree.insert('', 'end', values=(idx, row[0], row[1], row[2], row[3], row[4], row[5]))

    def delete_record(self):
        val = self.rec_var.get().strip()
        if not val.isdigit():
            messagebox.showwarning('Input Error', 'Please enter a valid numeric Record Number')
            return
        recno = int(val)
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) FROM assignments WHERE RecordNo = ?', (recno,))
        exists = cur.fetchone()[0]
        if not exists:
            conn.close()
            messagebox.showinfo('Not Found', 'No record with that Record Number')
            return
        if messagebox.askyesno('Confirm Delete', f'Are you sure you want to delete RecordNo {recno}?'):
            cur.execute('DELETE FROM assignments WHERE RecordNo = ?', (recno,))
            conn.commit()
            conn.close()
            messagebox.showinfo('Deleted', f'Record {recno} deleted')
            self.rec_var.set('')
            self.refresh_preview()
        else:
            conn.close()


# ---------------------------
# Run app
# ---------------------------
if __name__ == '__main__':
    try:
        from PIL import Image, ImageTk
    except Exception:
        # PIL optional, logos will be skipped if not present
        pass
    app = IGNOUApp()
    app.mainloop()
